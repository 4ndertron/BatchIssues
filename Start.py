import openpyxl as xl
from os import path
from os import walk
import re

issues = re.compile(r'Batch \d* Issues')
batch_file = re.compile(r'Batch \d*?')
batch_number = re.compile(r'\d+')

file = 'C:\\users\\robert.anderson\\Downloads\\Batches\\Batch 277 Funding  Request Details.xlsx'

issues_wb = xl.load_workbook(path.join(path.split(file)[0], 'Issues Iterated.xlsx'))
issues_page = issues_wb.active
issues_page.delete_rows(2, issues_page.max_row)

for folder_name, sub_folders, file_names in walk(path.split(file)[0]):
    print('The current folder is ' + folder_name)
    for sub_folder in sub_folders:
        print('SUBFOLDER OF ' + folder_name + ': ' + sub_folder)
    for file_name in file_names:
        print('FILE INSIDE ' + folder_name + ': ' + file_name)
        if batch_file.search(file_name):
            print('Opening batch file...')
            wb = xl.load_workbook(path.join(path.split(file)[0], file_name))
            print('Searching for issue page...')
            for i in range(len(wb.sheetnames)):
                if issues.search(wb.sheetnames[i]):
                    print('The issues page is found on sheet index ' + str(i) + ' of the current file')
                    print('The name of the sheet is ' + wb.sheetnames[i])
                    print('Beginning copy...')
                    ld_issues = wb[wb.sheetnames[i]]
                    for row in ld_issues.iter_rows(min_row=2, max_row=ld_issues.max_row, max_col=ld_issues.max_column,
                                                   values_only=True):
                        new_row = []
                        this_batch = batch_number.search(wb.sheetnames[i])
                        for value in row:
                            new_row.append(value)
                        new_row.append(this_batch.group())
                        issues_page.append(new_row)
                    print('Copy complete.')

for row in issues_page.values:
    print(row)

issues_wb.save(path.join(path.split(file)[0], 'Issues Iterated.xlsx'))
